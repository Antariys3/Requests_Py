import openpyxl
from Парсирг_инет_магаз import array


def writer(parameters):
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Товар"
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 50

    row = 1
    column = 1
    for item in parameters():
        cell = sheet.cell(row=row, column=column, value=item[0])
        cell_2 = sheet.cell(row=row, column=column + 1, value=item[1])
        cell_3 = sheet.cell(row=row, column=column + 2, value=item[2])
        print("Сохранено рядок", row)
        row += 1

    book.save("Товары с магазина.xlsx")


if __name__ == "__main__":
    writer(array)
